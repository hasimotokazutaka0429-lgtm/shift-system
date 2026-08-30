
import sqlite3
import streamlit as st

import calendar


from ortools.sat.python import cp_model

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from io import BytesIO


# ============================================================
# SQLite データベース
# ============================================================

DB_FILE = "shift_system.db"


def init_database():
    """データベースを初期化する"""

    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()


    # ========================================================
    # 社員テーブル
    # ========================================================

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS employees (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            employment_type TEXT NOT NULL,
            gender TEXT NOT NULL,
            experience INTEGER NOT NULL,
            leader INTEGER NOT NULL,
            day_min INTEGER NOT NULL,
            day_max INTEGER NOT NULL,
            night_min INTEGER NOT NULL,
            night_max INTEGER NOT NULL
        )
    """)


    # ========================================================
    # 勤務希望テーブル
    # ========================================================

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS requests (
            id INTEGER PRIMARY KEY AUTOINCREMENT,

            employee_id INTEGER NOT NULL,

            year INTEGER,

            month INTEGER NOT NULL,

            day INTEGER NOT NULL,

            request_type TEXT NOT NULL,

            FOREIGN KEY (employee_id)
                REFERENCES employees(id)
        )
    """)


    # ========================================================
    # 既存データベースへの対応
    # ========================================================

    cursor.execute("""
        PRAGMA table_info(requests)
    """)

    columns = [
        row[1]
        for row in cursor.fetchall()
    ]


    # 以前のデータベースにyearがなかった場合
    if "year" not in columns:

        cursor.execute("""
            ALTER TABLE requests
            ADD COLUMN year INTEGER
        """)

        # 以前のデータは2026年として扱う
        cursor.execute("""
            UPDATE requests
            SET year = 2026
            WHERE year IS NULL
        """)


    conn.commit()
    conn.close()




def get_employees():
    """データベースから社員情報を取得する"""

    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row

    cursor = conn.cursor()

    cursor.execute("""
        SELECT *
        FROM employees
        ORDER BY id
    """)

    employees = [
        dict(row)
        for row in cursor.fetchall()
    ]

    conn.close()

    return employees


def add_employee(
    name,
    employment_type,
    gender,
    experience,
    leader,
    day_min,
    day_max,
    night_min,
    night_max
):
    """社員をデータベースに登録する"""

    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()

    cursor.execute("""
        INSERT INTO employees (
            name,
            employment_type,
            gender,
            experience,
            leader,
            day_min,
            day_max,
            night_min,
            night_max
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        name,
        employment_type,
        gender,
        experience,
        leader,
        day_min,
        day_max,
        night_min,
        night_max
    ))

    conn.commit()
    conn.close()


def delete_employee(employee_id):
    """社員を削除する"""

    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()

    cursor.execute(
        "DELETE FROM employees WHERE id = ?",
        (employee_id,)
    )



    conn.commit()
    conn.close()



def update_employee(
    employee_id,
    name,
    employment_type,
    gender,
    experience,
    leader,
    day_min,
    day_max,
    night_min,
    night_max
):
    """社員情報を更新する"""

    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()

    cursor.execute("""
        UPDATE employees
        SET
            name = ?,
            employment_type = ?,
            gender = ?,
            experience = ?,
            leader = ?,
            day_min = ?,
            day_max = ?,
            night_min = ?,
            night_max = ?
        WHERE id = ?
    """, (
        name,
        employment_type,
        gender,
        experience,
        leader,
        day_min,
        day_max,
        night_min,
        night_max,
        employee_id
    ))

    conn.commit()
    conn.close()



def save_request(
    employee_id,
    year,
    month,
    day,
    request_type
):
    """勤務希望を保存する"""

    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()


    # 同じ社員・同じ年月日について
    # 既存の希望を削除
    cursor.execute("""
        DELETE FROM requests
        WHERE employee_id = ?
        AND year = ?
        AND month = ?
        AND day = ?
    """, (
        employee_id,
        year,
        month,
        day
    ))


    # 新しい希望を登録
    cursor.execute("""
        INSERT INTO requests (
            employee_id,
            year,
            month,
            day,
            request_type
        )
        VALUES (?, ?, ?, ?, ?)
    """, (
        employee_id,
        year,
        month,
        day,
        request_type
    ))


    conn.commit()
    conn.close()





def get_requests(
    employee_id,
    year,
    month
):
    """指定した社員・年月の勤務希望を取得する"""

    conn = sqlite3.connect(DB_FILE)

    conn.row_factory = sqlite3.Row

    cursor = conn.cursor()


    cursor.execute("""
        SELECT *
        FROM requests
        WHERE employee_id = ?
        AND year = ?
        AND month = ?
        ORDER BY day
    """, (
        employee_id,
        year,
        month
    ))


    requests = [
        dict(row)
        for row in cursor.fetchall()
    ]


    conn.close()

    return requests





def delete_request(
    employee_id,
    year,
    month,
    day
):
    """指定した勤務希望を削除する"""

    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()


    cursor.execute("""
        DELETE FROM requests
        WHERE employee_id = ?
        AND year = ?
        AND month = ?
        AND day = ?
    """, (
        employee_id,
        year,
        month,
        day
    ))


    conn.commit()
    conn.close()







# ============================================================
# 基本設定
# ============================================================

DAY_SHIFT = 0
NIGHT_SHIFT = 1
OFF_SHIFT = 2

SHIFT_NAMES = {
    DAY_SHIFT: "日勤",
    NIGHT_SHIFT: "夜勤",
    OFF_SHIFT: "休み"
}

init_database()
# ============================================================
# ページ設定
# ============================================================


st.set_page_config(
    page_title="シフト自動生成システム",
    layout="wide"
)

st.title("シフト自動生成システム")

st.write(
    "社員情報・勤務条件・希望を入力すると、"
    "OR-Toolsを使って30日分のシフトを自動生成します。"
)


# ============================================================
# セッションに社員データを保存
# ============================================================

if "employees" not in st.session_state:
    st.session_state.employees = []


# ============================================================
# 基本条件
# ============================================================

st.header("1. 基本条件")

col1, col2 = st.columns(2)

with col1:
    days = st.number_input(
        "シフト日数",
        min_value=1,
        max_value=31,
        value=30
    )

    required_day = st.number_input(
        "1日に必要な日勤人数",
        min_value=0,
        value=2
    )

    required_night = st.number_input(
        "1日に必要な夜勤人数",
        min_value=0,
        value=1
    )


with col2:

    required_full_time = st.number_input(
        "1日に必要な正社員人数",
        min_value=0,
        value=1
    )

    required_male = st.number_input(
        "1日に必要な男性人数",
        min_value=0,
        value=1
    )

    required_female = st.number_input(
        "1日に必要な女性人数",
        min_value=0,
        value=1
    )


col3, col4 = st.columns(2)

with col3:

    experienced_year = st.number_input(
        "経験者と判断する経験年数",
        min_value=0,
        value=3
    )

    required_experienced = st.number_input(
        "1日に必要な経験者人数",
        min_value=0,
        value=1
    )


with col4:

    required_day_leader = st.number_input(
        "日勤に必要なリーダー人数",
        min_value=0,
        value=1
    )

    required_night_leader = st.number_input(
        "夜勤に必要なリーダー人数",
        min_value=0,
        value=1
    )

    max_consecutive = st.number_input(
        "最大連勤日数",
        min_value=1,
        max_value=30,
        value=5
    )


# ============================================================
# 勤務ルール
# ============================================================

st.header("2. 勤務ルール")

night_next_off = st.checkbox(
    "夜勤の翌日は休みにする",
    value=True
)


# ============================================================
# 社員登録
# ============================================================

st.header("社員登録")


name = st.text_input(
    "氏名"
)


employment_type = st.selectbox(
    "雇用区分",
    ["正社員", "準社員"]
)


gender = st.selectbox(
    "性別",
    ["男性", "女性", "その他"]
)


experience = st.number_input(
    "経験年数",
    min_value=0,
    max_value=50,
    value=1
)


leader = st.checkbox(
    "リーダーができる"
)


st.subheader("日勤回数")

day_min = st.number_input(
    "日勤 最低回数",
    min_value=0,
    max_value=30,
    value=0
)

day_max = st.number_input(
    "日勤 最大回数",
    min_value=0,
    max_value=30,
    value=30
)


st.subheader("夜勤回数")

night_min = st.number_input(
    "夜勤 最低回数",
    min_value=0,
    max_value=30,
    value=0
)

night_max = st.number_input(
    "夜勤 最大回数",
    min_value=0,
    max_value=30,
    value=30
)


if st.button("社員を登録"):

    if name.strip() == "":
        st.error("氏名を入力してください。")

    elif day_min > day_max:
        st.error(
            "日勤の最低回数は最大回数以下にしてください。"
        )

    elif night_min > night_max:
        st.error(
            "夜勤の最低回数は最大回数以下にしてください。"
        )

    else:

        add_employee(
            name=name,
            employment_type=employment_type,
            gender=gender,
            experience=experience,
            leader=int(leader),
            day_min=day_min,
            day_max=day_max,
            night_min=night_min,
            night_max=night_max
        )

        st.success(
            f"{name}さんを登録しました。"
        )

        st.rerun()




# ============================================================
# 登録済み社員
# ============================================================

st.header("登録済み社員")


employees = get_employees()


if len(employees) == 0:

    st.info(
        "まだ社員が登録されていません。"
    )

else:

    for employee in employees:

        col1, col2 = st.columns([5, 1])

        with col1:

            leader_text = (
                "○ リーダー可"
                if employee["leader"]
                else "× リーダー不可"
            )

            st.write(
                f"**{employee['name']}** "
                f"｜{employee['employment_type']} "
                f"｜{employee['gender']} "
                f"｜経験{employee['experience']}年 "
                f"｜{leader_text}"
            )

            st.write(
                f"日勤："
                f"{employee['day_min']}～"
                f"{employee['day_max']}回　"
                f""
                f"夜勤："
                f"{employee['night_min']}～"
                f"{employee['night_max']}回"
            )

        with col2:

            if st.button(
                "削除",
                key=f"delete_{employee['id']}"
            ):

                delete_employee(
                    employee["id"]
                )

                st.success(
                    f"{employee['name']}さんを削除しました。"
                )

                st.rerun()



# ============================================================
# 社員情報の編集
# ============================================================

st.header("社員情報の編集")


employees = get_employees()


if len(employees) == 0:

    st.info(
        "編集できる社員が登録されていません。"
    )

else:

    # --------------------------------------------------------
    # 編集する社員を選択
    # --------------------------------------------------------

    selected_id = st.selectbox(
        "編集する社員",
        [employee["id"] for employee in employees],
        format_func=lambda employee_id:
            next(
                employee["name"]
                for employee in employees
                if employee["id"] == employee_id
            )
    )


    # 選択された社員を取得
    employee = next(
        employee
        for employee in employees
        if employee["id"] == selected_id
    )


    # --------------------------------------------------------
    # 現在の情報を入力欄に表示
    # --------------------------------------------------------

    edit_name = st.text_input(
        "氏名",
        value=employee["name"],
        key=f"edit_name_{selected_id}"
    )


    employment_types = [
        "正社員",
        "準社員"
    ]

    edit_employment_type = st.selectbox(
        "雇用区分",
        employment_types,
        index=employment_types.index(
            employee["employment_type"]
        ),
        key=f"edit_employment_{selected_id}"
    )


    genders = [
        "男性",
        "女性",
        "その他"
    ]

    edit_gender = st.selectbox(
        "性別",
        genders,
        index=genders.index(
            employee["gender"]
        ),
        key=f"edit_gender_{selected_id}"
    )


    edit_experience = st.number_input(
        "経験年数",
        min_value=0,
        max_value=50,
        value=employee["experience"],
        key=f"edit_experience_{selected_id}"
    )


    edit_leader = st.checkbox(
        "リーダーができる",
        value=bool(employee["leader"]),
        key=f"edit_leader_{selected_id}"
    )


    st.subheader("日勤回数")


    edit_day_min = st.number_input(
        "日勤 最低回数",
        min_value=0,
        max_value=30,
        value=employee["day_min"],
        key=f"edit_day_min_{selected_id}"
    )


    edit_day_max = st.number_input(
        "日勤 最大回数",
        min_value=0,
        max_value=30,
        value=employee["day_max"],
        key=f"edit_day_max_{selected_id}"
    )


    st.subheader("夜勤回数")


    edit_night_min = st.number_input(
        "夜勤 最低回数",
        min_value=0,
        max_value=30,
        value=employee["night_min"],
        key=f"edit_night_min_{selected_id}"
    )


    edit_night_max = st.number_input(
        "夜勤 最大回数",
        min_value=0,
        max_value=30,
        value=employee["night_max"],
        key=f"edit_night_max_{selected_id}"
    )


    # --------------------------------------------------------
    # 更新ボタン
    # --------------------------------------------------------

    if st.button(
        "社員情報を更新",
        key=f"update_{selected_id}"
    ):

        if edit_name.strip() == "":
            st.error(
                "氏名を入力してください。"
            )

        elif edit_day_min > edit_day_max:
            st.error(
                "日勤の最低回数は最大回数以下にしてください。"
            )

        elif edit_night_min > edit_night_max:
            st.error(
                "夜勤の最低回数は最大回数以下にしてください。"
            )

        else:

            update_employee(
                employee_id=selected_id,
                name=edit_name,
                employment_type=edit_employment_type,
                gender=edit_gender,
                experience=edit_experience,
                leader=int(edit_leader),
                day_min=edit_day_min,
                day_max=edit_day_max,
                night_min=edit_night_min,
                night_max=edit_night_max
            )

            st.success(
                f"{edit_name}さんの情報を更新しました。"
            )

            st.rerun()



     

# ============================================================
# 希望休・希望勤務
# ============================================================

st.header("希望休・希望勤務")

employees = get_employees()


if len(employees) == 0:

    st.info(
        "先に社員を登録してください。"
    )

else:

    # ========================================================
    # 年・月の指定
    # ========================================================

    col1, col2 = st.columns(2)

    with col1:

        selected_year = st.number_input(
            "対象年",
            min_value=2026,
            max_value=2100,
            value=2026,
            step=1
        )

    with col2:

        selected_month = st.number_input(
            "対象月",
            min_value=1,
            max_value=12,
            value=9,
            step=1
        )


    # ========================================================
    # 月の日数を取得
    # ========================================================

    days_in_month = calendar.monthrange(
        selected_year,
        selected_month
    )[1]


    st.subheader(
        f"{selected_year}年{selected_month}月の勤務希望"
    )


    st.write(
        "「指定なし」は希望なし、「休み・日勤・夜勤」は希望を表します。"
    )


    # ========================================================
    # 希望の選択肢
    # ========================================================

    request_options = [
        "指定なし",
        "休み",
        "日勤",
        "夜勤"
    ]


    # ========================================================
    # 全社員の希望を取得
    # ========================================================

    all_requests = {}


    for employee in employees:

        employee_id = employee["id"]

        requests = get_requests(
            employee_id,
            selected_year,
            selected_month
        )


        all_requests[employee_id] = {
            request["day"]: request["request_type"]
            for request in requests
        }


    # ========================================================
    # 横スクロール用
    # ========================================================

    st.markdown(
        """
        <style>

        .shift-table-container {
            overflow-x: auto;
            width: 100%;
            border: 1px solid #ddd;
            padding: 10px;
        }

        .shift-table {
            min-width: 1800px;
        }

        </style>
        """,
        unsafe_allow_html=True
    )


    # ========================================================
    # 表のヘッダー
    # ========================================================

    st.markdown(
        '<div class="shift-table-container">',
        unsafe_allow_html=True
    )

    st.markdown(
        '<div class="shift-table">',
        unsafe_allow_html=True
    )


    # --------------------------------------------------------
    # 社員名 + 日付
    # --------------------------------------------------------

    header_cols = st.columns(
        [2] + [1] * days_in_month
    )


    with header_cols[0]:

        st.markdown(
            "**社員名**"
        )


    for day in range(1, days_in_month + 1):

        weekday = calendar.weekday(
            selected_year,
            selected_month,
            day
        )


        weekday_names = [
            "月",
            "火",
            "水",
            "木",
            "金",
            "土",
            "日"
        ]


        with header_cols[day]:

            st.markdown(
                f"**{day}日**"
            )

            st.caption(
                f"({weekday_names[weekday]})"
            )


    st.divider()


    # ========================================================
    # 社員ごとの希望入力
    # ========================================================

    selected_requests = {}


    for employee in employees:

        employee_id = employee["id"]

        employee_name = employee["name"]


        # 社員ごとの辞書
        selected_requests[employee_id] = {}


        cols = st.columns(
            [2] + [1] * days_in_month
        )


        # ----------------------------------------------------
        # 社員名
        # ----------------------------------------------------

        with cols[0]:

            st.markdown(
                f"**{employee_name}**"
            )


        # ----------------------------------------------------
        # 1日～月末
        # ----------------------------------------------------

        for day in range(
            1,
            days_in_month + 1
        ):

            current_request = all_requests[
                employee_id
            ].get(
                day,
                "指定なし"
            )


            with cols[day]:

                selected_requests[
                    employee_id
                ][day] = st.selectbox(
                    "希望",
                    request_options,
                    index=request_options.index(
                        current_request
                    ),
                    key=(
                        f"request_"
                        f"{employee_id}_"
                        f"{selected_year}_"
                        f"{selected_month}_"
                        f"{day}"
                    ),
                    label_visibility="collapsed"
                )


    st.markdown(
        "</div></div>",
        unsafe_allow_html=True
    )


    # ========================================================
    # 保存ボタン
    # ========================================================

    st.write("")


    if st.button(
        "全社員の希望を保存",
        type="primary"
    ):

        for employee in employees:

            employee_id = employee["id"]


            for day in range(
                1,
                days_in_month + 1
            ):

                request = selected_requests[
                    employee_id
                ][day]


                # --------------------------------------------
                # 指定なし
                # --------------------------------------------

                if request == "指定なし":

                    delete_request(
                        employee_id=employee_id,
                        year=selected_year,
                        month=selected_month,
                        day=day
                    )


                # --------------------------------------------
                # 希望あり
                # --------------------------------------------

                else:

                    save_request(
                        employee_id=employee_id,
                        year=selected_year,
                        month=selected_month,
                        day=day,
                        request_type=request
                    )


        st.success(
            f"{selected_year}年{selected_month}月の希望を保存しました。"
        )

        st.rerun()



# ============================================================
# シフト生成関数
# ============================================================

def generate_shift(
    employees,
    days,
    required_day,
    required_night,
    required_full_time,
    required_male,
    required_female,
    experienced_year,
    required_experienced,
    required_day_leader,
    required_night_leader,
    max_consecutive,
    night_next_off
):

    model = cp_model.CpModel()

    employee_count = len(employees)


    # ========================================================
    # 変数
    # ========================================================

    work = {}

    for e in range(employee_count):

        for d in range(days):

            for s in range(3):

                work[e, d, s] = model.NewBoolVar(
                    f"work_{e}_{d}_{s}"
                )


    # ========================================================
    # 1人1日1シフト
    # ========================================================

    for e in range(employee_count):

        for d in range(days):

            model.Add(
                sum(
                    work[e, d, s]
                    for s in range(3)
                )
                == 1
            )


    # ========================================================
    # 日勤人数
    # ========================================================

    for d in range(days):

        model.Add(
            sum(
                work[e, d, DAY_SHIFT]
                for e in range(employee_count)
            )
            == required_day
        )


    # ========================================================
    # 夜勤人数
    # ========================================================

    for d in range(days):

        model.Add(
            sum(
                work[e, d, NIGHT_SHIFT]
                for e in range(employee_count)
            )
            == required_night
        )


    # ========================================================
    # 正社員人数
    # ========================================================

    for d in range(days):

        model.Add(
            sum(

                work[e, d, DAY_SHIFT]
                +
                work[e, d, NIGHT_SHIFT]

                for e in range(employee_count)

                if employees[e]["type"] == "正社員"

            )
            >= required_full_time
        )


    # ========================================================
    # 男性人数
    # ========================================================

    for d in range(days):

        model.Add(
            sum(

                work[e, d, DAY_SHIFT]
                +
                work[e, d, NIGHT_SHIFT]

                for e in range(employee_count)

                if employees[e]["gender"] == "男性"

            )
            >= required_male
        )


    # ========================================================
    # 女性人数
    # ========================================================

    for d in range(days):

        model.Add(
            sum(

                work[e, d, DAY_SHIFT]
                +
                work[e, d, NIGHT_SHIFT]

                for e in range(employee_count)

                if employees[e]["gender"] == "女性"

            )
            >= required_female
        )


    # ========================================================
    # 経験者人数
    # ========================================================

    for d in range(days):

        model.Add(
            sum(

                work[e, d, DAY_SHIFT]
                +
                work[e, d, NIGHT_SHIFT]

                for e in range(employee_count)

                if employees[e]["experience"]
                >= experienced_year

            )
            >= required_experienced
        )


    # ========================================================
    # 日勤リーダー
    # ========================================================

    for d in range(days):

        model.Add(
            sum(

                work[e, d, DAY_SHIFT]

                for e in range(employee_count)

                if employees[e]["leader"]

            )
            >= required_day_leader
        )


    # ========================================================
    # 夜勤リーダー
    # ========================================================

    for d in range(days):

        model.Add(
            sum(

                work[e, d, NIGHT_SHIFT]

                for e in range(employee_count)

                if employees[e]["leader"]

            )
            >= required_night_leader
        )


    # ========================================================
    # 個人の日勤回数
    # ========================================================

    for e in range(employee_count):

        model.Add(
            sum(
                work[e, d, DAY_SHIFT]
                for d in range(days)
            )
            >= employees[e]["min_day"]
        )

        model.Add(
            sum(
                work[e, d, DAY_SHIFT]
                for d in range(days)
            )
            <= employees[e]["max_day"]
        )


    # ========================================================
    # 個人の夜勤回数
    # ========================================================

    for e in range(employee_count):

        model.Add(
            sum(
                work[e, d, NIGHT_SHIFT]
                for d in range(days)
            )
            >= employees[e]["min_night"]
        )

        model.Add(
            sum(
                work[e, d, NIGHT_SHIFT]
                for d in range(days)
            )
            <= employees[e]["max_night"]
        )


    # ========================================================
    # 夜勤の翌日は休み
    # ========================================================

    if night_next_off:

        for e in range(employee_count):

            for d in range(days - 1):

                model.Add(
                    work[e, d, NIGHT_SHIFT]
                    <=
                    work[e, d + 1, OFF_SHIFT]
                )


    # ========================================================
    # 最大連勤日数
    # ========================================================

    for e in range(employee_count):

        for start in range(
            days - max_consecutive
        ):

            model.Add(

                sum(

                    work[e, d, DAY_SHIFT]
                    +
                    work[e, d, NIGHT_SHIFT]

                    for d in range(
                        start,
                        start + max_consecutive + 1
                    )

                )
                <= max_consecutive

            )


    # ========================================================
    # 希望を目的関数にする
    # ========================================================

    request_scores = []

    for e in range(employee_count):

        for d, requested_shift in (
            employees[e]["requests"].items()
        ):

            if requested_shift == "日勤":

                request_scores.append(
                    work[e, d, DAY_SHIFT]
                )

            elif requested_shift == "夜勤":

                request_scores.append(
                    work[e, d, NIGHT_SHIFT]
                )

            elif requested_shift == "休み":

                request_scores.append(
                    work[e, d, OFF_SHIFT]
                )


    if request_scores:

        model.Maximize(
            sum(request_scores)
        )


    # ========================================================
    # ソルバー
    # ========================================================

    solver = cp_model.CpSolver()

    solver.parameters.max_time_in_seconds = 30

    status = solver.Solve(model)


    if status not in [
        cp_model.OPTIMAL,
        cp_model.FEASIBLE
    ]:

        return None


    # ========================================================
    # 結果を辞書にする
    # ========================================================

    result = []

    for e in range(employee_count):

        employee_result = []

        for d in range(days):

            if solver.Value(
                work[e, d, DAY_SHIFT]
            ):

                employee_result.append("日勤")

            elif solver.Value(
                work[e, d, NIGHT_SHIFT]
            ):

                employee_result.append("夜勤")

            else:

                employee_result.append("休み")


        result.append(
            employee_result
        )


    return result


# ============================================================
# Excel作成関数
# ============================================================

def create_excel(
    employees,
    result,
    days
):

    wb = Workbook()

    ws = wb.active

    ws.title = "シフト表"


    # ========================================================
    # タイトル
    # ========================================================

    ws["A1"] = "30日間シフト表"

    ws["A1"].font = Font(
        size=16,
        bold=True
    )


    # ========================================================
    # ヘッダー
    # ========================================================

    headers = [
        "名前",
        "雇用形態",
        "性別",
        "経験年数",
        "リーダー"
    ]


    for col, header in enumerate(
        headers,
        start=1
    ):

        cell = ws.cell(
            row=3,
            column=col
        )

        cell.value = header

        cell.font = Font(
            bold=True
        )

        cell.alignment = Alignment(
            horizontal="center"
        )


    start_shift_column = 6


    for d in range(days):

        cell = ws.cell(
            row=3,
            column=start_shift_column + d
        )

        cell.value = f"{d + 1}日"

        cell.font = Font(
            bold=True
        )

        cell.alignment = Alignment(
            horizontal="center"
        )


    # ========================================================
    # 社員とシフト
    # ========================================================

    for e, employee in enumerate(
        employees
    ):

        row = e + 4


        ws.cell(
            row=row,
            column=1
        ).value = employee["name"]


        ws.cell(
            row=row,
            column=2
        ).value = employee["type"]


        ws.cell(
            row=row,
            column=3
        ).value = employee["gender"]


        ws.cell(
            row=row,
            column=4
        ).value = employee["experience"]


        ws.cell(
            row=row,
            column=5
        ).value = (
            "○"
            if employee["leader"]
            else "×"
        )


        for d in range(days):

            cell = ws.cell(
                row=row,
                column=start_shift_column + d
            )

            cell.value = result[e][d]

            cell.alignment = Alignment(
                horizontal="center"
            )


    # ========================================================
    # 枠線
    # ========================================================

    thin_border = Border(

        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")

    )


    for row in ws.iter_rows(

        min_row=3,
        max_row=3 + len(employees),
        min_col=1,
        max_col=start_shift_column + days - 1

    ):

        for cell in row:

            cell.border = thin_border

            cell.alignment = Alignment(
                horizontal="center"
            )


    # ========================================================
    # 列幅
    # ========================================================

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10


    for col in range(
        start_shift_column,
        start_shift_column + days
    ):

        letter = get_column_letter(col)

        ws.column_dimensions[
            letter
        ].width = 8


    # ========================================================
    # ウィンドウ固定
    # ========================================================

    ws.freeze_panes = "F4"


    # ========================================================
    # メモリ上にExcelを保存
    # ========================================================

    output = BytesIO()

    wb.save(output)

    output.seek(0)

    return output


# ============================================================
# シフト生成
# ============================================================

st.header("5. シフト生成")


if st.button(
    "シフトを自動生成",
    type="primary"
):

    if len(st.session_state.employees) == 0:

        st.error(
            "社員を1人以上登録してください。"
        )

    else:

        with st.spinner(
            "シフトを計算しています..."
        ):

            result = generate_shift(

                st.session_state.employees,

                days,

                required_day,
                required_night,

                required_full_time,

                required_male,
                required_female,

                experienced_year,
                required_experienced,

                required_day_leader,
                required_night_leader,

                max_consecutive,

                night_next_off

            )


        if result is None:

            st.error(
                "条件をすべて満たすシフトを"
                "作成できませんでした。"
            )

            st.info(
                "必要人数、社員数、勤務回数、"
                "リーダー条件、最大連勤日数などを"
                "確認してください。"
            )

        else:

            st.success(
                "シフトを生成しました！"
            )


            # ==================================================
            # 結果表示
            # ==================================================

            st.subheader("生成されたシフト")


            table_data = []


            for e, employee in enumerate(
                st.session_state.employees
            ):

                row = {

                    "名前":
                        employee["name"],

                    "雇用形態":
                        employee["type"],

                    "性別":
                        employee["gender"],

                    "経験年数":
                        employee["experience"],

                    "リーダー":
                        "○"
                        if employee["leader"]
                        else "×"

                }


                for d in range(days):

                    row[
                        f"{d + 1}日"
                    ] = result[e][d]


                table_data.append(row)


            st.dataframe(
                table_data,
                use_container_width=True
            )


            # ==================================================
            # Excel
            # ==================================================

            excel_file = create_excel(
                st.session_state.employees,
                result,
                days
            )


            st.download_button(

                label="Excelをダウンロード",

                data=excel_file,

                file_name="shift_result.xlsx",

                mime=(
                    "application/vnd.openxmlformats-officedocument."
                    "spreadsheetml.sheet"
                )

            )

